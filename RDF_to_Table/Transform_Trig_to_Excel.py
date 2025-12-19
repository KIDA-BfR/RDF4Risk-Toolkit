import argparse
import sys
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import defaultdict
from urllib.parse import urlparse

try:
    from rdflib import Dataset, URIRef, Literal
    from rdflib.namespace import RDF, RDFS, SKOS
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from tqdm import tqdm
except ImportError as e:
    print(f"Error: Missing required library. Please install dependencies:")
    print("  pip install rdflib pandas openpyxl tqdm")
    print(f"\nDetails: {e}")
    sys.exit(1)



# Named graph suffixes
DCAT_METADATA_SUFFIX = '/graph/dcat-metadata'
PUBLICATION_REFERENCE_SUFFIX = '/graph/publication-reference'

# Excel styling
HEADER_BG_COLOR = "366092"  # Blue
HEADER_FONT_COLOR = "FFFFFF"  # White
HYPERLINK_COLOR = "0563C1"  # Excel blue
MAX_COLUMN_WIDTH = 50
SAMPLE_ROWS_FOR_WIDTH = 100

# Excel formatting constants
EXCEL_HYPERLINK_MAX_LENGTH = 255
EXCEL_CELL_CHAR_PREVIEW = 100


class TriGConverter:
    """
    Converts TriG (RDF) files to Excel and Markdown formats.

    This class handles:
    - Parsing TriG files with named graphs
    - Extracting property labels and URI mappings
    - Building subject data with multi-valued properties
    - Expanding multi-valued properties into separate rows
    - Generating Excel files with HYPERLINK formulas
    - Creating comprehensive property mapping sheets
    """

    def __init__(self, input_file: Optional[Path] = None, data: Optional[str] = None):
        """
        Initialize the converter.

        Args:
            input_file: Path to the input TriG file
            data: Raw TriG data as string
        """
        self.input_file = input_file
        self.data = data
        self.graph: Optional[Dataset] = None

        # Property mappings
        self.property_labels: Dict[str, str] = {}  # URI -> Label
        self.label_to_property_uri: Dict[str, str] = {}  # Label -> URI

        # Data storage
        self.subjects_data: List[Dict] = []
        self.namespaces: Dict[str, str] = {}

        # URI mappings for hyperlinks
        self.uri_to_label: Dict[str, str] = {}  # All URIs to their labels
        self.uri_to_exact_match: Dict[str, str] = {}  # skos:exactMatch mappings
        self.uri_to_close_match: Dict[str, str] = {}  # skos:closeMatch mappings

        # Named graph data
        self.named_graphs_data: Dict[str, List[Tuple]] = defaultdict(list)

        # Performance caches
        self._property_counts_cache: Optional[Dict[str, int]] = None
        self._uri_validation_cache: Dict[str, bool] = {}  # Cache for validated URIs



    def _get_property_counts(self) -> Dict[str, int]:
        """
        Get property usage counts (cached for performance).

        Returns:
            Dictionary mapping property labels to their usage counts
        """
        if self._property_counts_cache is None:
            property_counts = defaultdict(int)
            for subject in self.subjects_data:
                for key in subject.keys():
                    if key != 'subject_uri':
                        property_counts[key] += 1
            self._property_counts_cache = dict(property_counts)

        return self._property_counts_cache

    def _extract_local_name(self, uri: str) -> str:
        """Extract the local name from a URI (last part after / or #)."""
        if '#' in uri:
            return uri.rsplit('#', 1)[1]
        elif '/' in uri:
            return uri.rsplit('/', 1)[1]
        return uri

    def _should_use_local_name(self, uri: str) -> bool:
        """Check if URI should use its local name and has no label."""
        # Check if it's an internal-looking URI (contains /concepts/ or /subjects/)
        is_internal = '/concepts/' in uri or '/subjects/' in uri
        return is_internal and uri not in self.uri_to_label

    def _is_internal_concept(self, uri: str) -> bool:
        """Check if URI is an internal concept (should not have hyperlinks)."""
        return '/concepts/' in uri

    def _get_hyperlink_uri(self, uri: str) -> str:
        """
        Get the best URI for hyperlinking.

        Preference order: exactMatch > closeMatch > original URI

        Args:
            uri: Original URI

        Returns:
            The best available URI for hyperlinking
        """
        if uri in self.uri_to_exact_match:
            return self.uri_to_exact_match[uri]
        elif uri in self.uri_to_close_match:
            return self.uri_to_close_match[uri]
        return uri



    def _validate_uri(self, uri: str) -> bool:
        """
        Validate a URI for use in Excel hyperlinks (cached for performance).

        Args:
            uri: The URI to validate

        Returns:
            True if URI is valid, False otherwise
        """
        # Check cache first
        if uri in self._uri_validation_cache:
            return self._uri_validation_cache[uri]

        # Validate inputs
        if not uri or not isinstance(uri, str):
            self._uri_validation_cache[uri] = False
            return False

        # Clean and validate URI
        uri = uri.strip()
        if not (uri.startswith('http://') or uri.startswith('https://')):
            self._uri_validation_cache[uri] = False
            return False

        # Validate length
        if len(uri) > EXCEL_HYPERLINK_MAX_LENGTH:
            self._uri_validation_cache[uri] = False
            return False

        # Check for invalid characters
        if re.search(r'\s', uri):  # No whitespace
            self._uri_validation_cache[uri] = False
            return False
        if re.search(r'[\x00-\x1F\x7F]', uri):  # No control characters
            self._uri_validation_cache[uri] = False
            return False

        # Validate URL structure
        try:
            parsed = urlparse(uri)
            if not parsed.scheme or not parsed.netloc:
                self._uri_validation_cache[uri] = False
                return False
        except Exception:
            self._uri_validation_cache[uri] = False
            return False

        self._uri_validation_cache[uri] = True
        return True

    def _safe_set_hyperlink_formula(self, cell, uri: str, display_value: str,
                                    override_font=None) -> bool:
        """
        Set a HYPERLINK formula on a cell (bypasses Excel's 65k native hyperlink limit).

        This method validates the URI thoroughly before creating the formula to
        prevent Excel corruption.

        Args:
            cell: openpyxl cell object
            uri: The URI to link to
            display_value: Display text for the link
            override_font: Optional Font object to use instead of default blue hyperlink style
                          (useful for headers where white text on blue background is needed)

        Returns:
            True if hyperlink was set successfully, False otherwise
        """
        # Validate inputs
        if not display_value or not isinstance(display_value, str):
            return False

        # Validate original URI
        if not self._validate_uri(uri):
            return False

        # Get the actual hyperlink URI (with exactMatch/closeMatch)
        hyperlink_uri = self._get_hyperlink_uri(uri)
        if not hyperlink_uri or not isinstance(hyperlink_uri, str):
            return False

        # Validate hyperlink URI
        if not self._validate_uri(hyperlink_uri):
            return False

        # Create and set HYPERLINK formula
        try:
            # Escape double quotes for Excel formula
            display_escaped = display_value.replace('"', '""')
            uri_escaped = hyperlink_uri.replace('"', '""')

            formula = f'=HYPERLINK("{uri_escaped}", "{display_escaped}")'
            cell.value = formula

            # Use override font if provided (e.g., for headers), otherwise use default blue style
            if override_font:
                cell.font = override_font
            else:
                cell.font = Font(color=HYPERLINK_COLOR, underline="single")
            return True
        except Exception:
            # Fallback to plain text
            cell.value = display_value
            return False



    def parse_trig(self) -> bool:
        """
        Parse the TriG file or data into an RDF graph.

        Returns:
            True if parsing succeeded, False otherwise
        """
        try:
            self.graph = Dataset()
            if self.input_file:
                print(f"Loading TriG file: {self.input_file}")
                self.graph.parse(self.input_file, format='trig')
            elif self.data:
                print("Loading TriG data from string")
                self.graph.parse(data=self.data, format='trig')
            else:
                print("Error: No input file or data provided")
                return False

            print(f"Successfully loaded {len(self.graph):,} triples")

            if len(self.graph) == 0 and self.input_file:
                print("Warning: No triples found. Trying alternative parsing...")
                with open(self.input_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.graph.parse(data=content, format='trig')
                print(f"Loaded {len(self.graph):,} triples with alternative method")

            return True
        except Exception as e:
            print(f"Error parsing TriG: {e}")
            import traceback
            traceback.print_exc()
            return False

    def extract_all_data(self):
        """
        Extract all data in a single optimized pass through the graph.

        This method:
        1. Collects all triples organized by subject and graph
        2. Extracts property labels, URI mappings, and match relationships
        3. Dynamically identifies data vs metadata graphs
        4. Builds subject data dictionaries
        """
        print("Extracting data from graph (optimized single-pass)...")

        # Data collection structures
        properties_with_labels = {}
        property_types = set()
        graph_triples = defaultdict(list)
        all_subjects = set()

        # Iterate through all quads in the dataset
        total_triples = len(self.graph)
        print(f"Processing {total_triples:,} quads...")

        # Single pass through all quads
        for s, p, o, g in tqdm(self.graph.quads((None, None, None, None)), 
                              total=total_triples, desc="Loading quads", unit="quad"):
            graph_uri = str(g) if g else None

            # Store triples by graph for later processing
            graph_triples[graph_uri].append((s, p, o))
            all_subjects.add(s)

            # Extract namespaces
            self._extract_namespace(s)
            self._extract_namespace(p)
            self._extract_namespace(o)

            # Extract labels (rdfs:label and skos:prefLabel)
            if p == RDFS.label or p == SKOS.prefLabel:
                uri = str(s)
                label = str(o)
                properties_with_labels[uri] = label
                # Prefer skos:prefLabel over rdfs:label
                if uri not in self.uri_to_label or p == SKOS.prefLabel:
                    self.uri_to_label[uri] = label

            # Extract skos:exactMatch relationships
            if p == SKOS.exactMatch:
                self.uri_to_exact_match[str(s)] = str(o)

            # Extract skos:closeMatch relationships
            if p == SKOS.closeMatch:
                self.uri_to_close_match[str(s)] = str(o)

            # Extract property type declarations
            if p == RDF.type and o == RDF.Property:
                property_types.add(str(s))

        # Build property labels and mappings
        for prop_uri in property_types:
            if prop_uri in properties_with_labels:
                label = properties_with_labels[prop_uri]
                self.property_labels[prop_uri] = label
                self.label_to_property_uri[label] = prop_uri

        # Identify graphs and extract subjects
        subjects_dict = defaultdict(list)
        
        # Categorize graphs
        metadata_graphs = []
        data_graphs = []
        
        for graph_uri, triples in graph_triples.items():
            if not graph_uri:
                data_graphs.append(graph_uri)
                continue
                
            if graph_uri.endswith(DCAT_METADATA_SUFFIX) or graph_uri.endswith(PUBLICATION_REFERENCE_SUFFIX):
                self.named_graphs_data[graph_uri] = triples
                metadata_graphs.append(graph_uri)
            else:
                data_graphs.append(graph_uri)
        
        # Identify main subjects from data graphs
        # A subject is a 'main subject' if it's in a data graph AND not a property
        for graph_uri in data_graphs:
            for s, p, o in graph_triples[graph_uri]:
                # Heuristic: focusing on row subjects (usually containing /subjects/)
                # and excluding properties and metadata-related subjects
                s_str = str(s)
                if s_str not in property_types and '/subjects/' in s_str:
                    subjects_dict[s].append((p, o))

        # Fallback: if no /subjects/ found, take all subjects from data graphs that aren't properties
        if not subjects_dict and data_graphs:
            for graph_uri in data_graphs:
                for s, p, o in graph_triples[graph_uri]:
                    if str(s) not in property_types:
                        subjects_dict[s].append((p, o))

        # Print statistics
        print(f"Found {len(self.property_labels)} property labels (declared as rdf:Property)")
        print(f"Found {len(self.namespaces)} namespaces")
        print(f"Found {len(subjects_dict)} data subjects across {len(data_graphs)} graphs")
        print(f"Found {len(metadata_graphs)} metadata graphs")
        print(f"Found {len(self.uri_to_exact_match)} exactMatch relationships")
        print(f"Found {len(self.uri_to_close_match)} closeMatch relationships")

        # Build subject data structures
        self._build_subject_data(subjects_dict)

    def _extract_namespace(self, uri):
        """Extract and store namespace from a URI."""
        if not isinstance(uri, URIRef):
            return

        uri_str = str(uri)
        if '#' in uri_str:
            ns = uri_str.rsplit('#', 1)[0] + '#'
        elif '/' in uri_str:
            parts = uri_str.rsplit('/', 1)
            if len(parts[1]) > 0:
                ns = parts[0] + '/'
            else:
                return
        else:
            return

        if ns not in self.namespaces:
            self.namespaces[ns] = ns

    def _build_subject_data(self, subjects_dict: Dict):
        """
        Build subject data structures from collected triples.

        Args:
            subjects_dict: Dictionary of subject URIs to their predicates and objects
        """
        print("Building subject data structures...")

        for subject in tqdm(sorted(subjects_dict.keys(), key=str),
                           desc="Processing subjects", unit="subject"):
            subject_data = {'subject_uri': str(subject)}

            for p, o in subjects_dict[subject]:
                if p == RDF.type:
                    continue

                prop_uri = str(p)
                prop_label = self.property_labels.get(prop_uri, prop_uri)

                # Populate reverse mapping for ALL properties used in data
                if prop_label != prop_uri and prop_label not in self.label_to_property_uri:
                    self.label_to_property_uri[prop_label] = prop_uri

                # Extract value
                value = str(o)

                # Handle multiple values for same property
                if prop_label in subject_data:
                    if not isinstance(subject_data[prop_label], list):
                        subject_data[prop_label] = [subject_data[prop_label]]
                    subject_data[prop_label].append(value)
                else:
                    subject_data[prop_label] = value

            self.subjects_data.append(subject_data)

        print(f"Extracted data for {len(self.subjects_data)} subjects")
        print(f"Total property label mappings: {len(self.label_to_property_uri)}")

    def expand_list_values(self):
        """
        Expand rows with list values into multiple rows.

        This enables proper hyperlinks for multi-valued properties in Excel.
        Each value in a list gets its own row, with other properties duplicated.
        """
        print("Expanding list values into separate rows...")

        expanded_data = []
        rows_expanded_count = 0

        for subject in tqdm(self.subjects_data, desc="Expanding rows", unit="subject"):
            # Separate list properties from scalar properties
            list_properties = {}
            scalar_properties = {}

            for key, value in subject.items():
                if isinstance(value, list) and len(value) > 0:
                    list_properties[key] = value
                else:
                    scalar_properties[key] = value

            # If no list properties, keep row as-is
            if not list_properties:
                expanded_data.append(subject)
                continue

            # Create one row per value (using max length of all lists)
            max_list_length = max(len(v) for v in list_properties.values())

            for i in range(max_list_length):
                new_row = scalar_properties.copy()

                for prop_key, prop_values in list_properties.items():
                    if i < len(prop_values):
                        new_row[prop_key] = prop_values[i]
                    else:
                        # For shorter lists, repeat the last value
                        new_row[prop_key] = prop_values[-1]

                expanded_data.append(new_row)

            rows_expanded_count += 1

        original_count = len(self.subjects_data)
        self.subjects_data = expanded_data

        print(f"Expanded {rows_expanded_count} subjects with list values")
        print(f"Rows before: {original_count:,}, after: {len(self.subjects_data):,}")



    def export_to_excel(self, output_file: Path):
        """
        Export data to Excel format with HYPERLINK formulas.

        This method creates:
        1. RDF Data sheet with all subjects and their properties
        2. Property Mappings sheet with property definitions

        Uses =HYPERLINK() formulas to bypass Excel's 65k hyperlink limit.
        """
        print(f"Exporting to Excel: {output_file}")

        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create RDF Data sheet
            self._create_rdf_data_sheet(wb)

            # Create Property Mappings sheet
            self._create_property_mappings_sheet(wb)

            # Save workbook
            wb.save(output_file)

            self._print_export_summary(output_file)

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            import traceback
            traceback.print_exc()

    def _create_rdf_data_sheet(self, wb: Workbook):
        """Create the main RDF Data sheet with subjects and properties."""
        ws_data = wb.create_sheet('RDF Data', 0)
        print("Creating RDF Data sheet with HYPERLINK formulas...")

        if not self.subjects_data:
            return

        # Collect all unique property keys (excluding subject_uri)
        all_keys = set()
        for subject in self.subjects_data:
            all_keys.update(subject.keys())
        all_keys.discard('subject_uri')
        headers = sorted(all_keys)

        # Write headers
        self._write_headers(ws_data, headers)

        # Write data rows
        self._write_data_rows(ws_data, headers)

        # Auto-adjust column widths
        self._adjust_column_widths(ws_data, headers)

    def _write_headers(self, ws, headers: List[str]):
        """Write column headers with proper formatting and hyperlinks."""
        header_fill = PatternFill(start_color=HEADER_BG_COLOR,
                                  end_color=HEADER_BG_COLOR, fill_type="solid")
        header_font_link = Font(bold=True, color=HEADER_FONT_COLOR, underline="single")
        header_font_plain = Font(bold=True, color=HEADER_FONT_COLOR)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill

            # Add hyperlink if property has URI with external match
            if header in self.label_to_property_uri:
                prop_uri = self.label_to_property_uri[header]

                # Only create hyperlinks if property has exactMatch or closeMatch
                # to an external resource (not for bare internal concept URIs)
                has_external_match = (prop_uri in self.uri_to_exact_match or
                                     prop_uri in self.uri_to_close_match)

                if has_external_match:
                    if self._safe_set_hyperlink_formula(cell, prop_uri, header,
                                                       override_font=header_font_link):
                        # Font already set by _safe_set_hyperlink_formula with override
                        pass
                    else:
                        # Hyperlink failed - use plain text
                        cell.value = header
                        cell.font = header_font_plain
                else:
                    # No external match - plain text
                    cell.value = header
                    cell.font = header_font_plain
            else:
                # No property URI (e.g., "Subject URI")
                cell.value = header
                cell.font = header_font_plain

    def _write_data_rows(self, ws, headers: List[str]):
        """Write data rows with hyperlinks for URIs."""
        for row_idx, subject in enumerate(tqdm(self.subjects_data,
                                               desc="Writing data rows", unit="row"), 2):
            # Write property values (Subject URI excluded)
            for col_idx, header in enumerate(headers, 1):
                self._write_property_cell(ws, row_idx, col_idx, subject, header)

    def _write_subject_uri_cell(self, ws, row_idx: int, subject: Dict):
        """Write the Subject URI cell with appropriate formatting."""
        subject_uri = subject.get('subject_uri', '')
        cell = ws.cell(row=row_idx, column=1)

        if subject_uri in self.uri_to_label:
            # Has label: show with hyperlink
            display_label = self.uri_to_label[subject_uri]
            if not self._safe_set_hyperlink_formula(cell, subject_uri, display_label):
                cell.value = display_label
        elif self._should_use_local_name(subject_uri):
            # Internal concept without label
            cell.value = self._extract_local_name(subject_uri)
        else:
            # Plain URI
            cell.value = subject_uri if subject_uri else ''

    def _write_property_cell(self, ws, row_idx: int, col_idx: int,
                            subject: Dict, header: str):
        """Write a property value cell with appropriate formatting."""
        value = subject.get(header, '')

        # Ensure proper None handling
        if value is None or value == '':
            value_str = ''
        else:
            value_str = str(value)

        cell = ws.cell(row=row_idx, column=col_idx)

        # Apply label and hyperlink if available
        if value_str and value_str in self.uri_to_label:
            display_label = self.uri_to_label[value_str]
            if not self._safe_set_hyperlink_formula(cell, value_str, display_label):
                cell.value = display_label
        elif value_str and self._should_use_local_name(value_str):
            cell.value = self._extract_local_name(value_str)
        else:
            cell.value = value_str

    def _adjust_column_widths(self, ws, headers: List[str]):
        """Auto-adjust column widths based on content."""
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))

            # Sample first N rows for width calculation
            for row_idx in range(2, min(SAMPLE_ROWS_FOR_WIDTH + 2,
                                       len(self.subjects_data) + 2)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length,
                                    len(str(cell.value)[:EXCEL_CELL_CHAR_PREVIEW]))

            adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    def _create_property_mappings_sheet(self, wb: Workbook):
        """Create the Property Mappings sheet."""
        ws_props = wb.create_sheet('Property Mappings', 1)
        print("Creating Property Mappings sheet...")

        # Write headers
        header_fill = PatternFill(start_color=HEADER_BG_COLOR,
                                  end_color=HEADER_BG_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=HEADER_FONT_COLOR)

        ws_props.cell(row=1, column=1).value = "Property URI"
        ws_props.cell(row=1, column=2).value = "Label"
        ws_props.cell(row=1, column=3).value = "Usage Count"

        for col in range(1, 4):
            cell = ws_props.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        # Get cached property usage counts
        property_counts = self._get_property_counts()

        # Write property data
        for row_idx, (uri, label) in enumerate(
            sorted(self.property_labels.items(), key=lambda x: x[1]), 2):

            cell_uri = ws_props.cell(row=row_idx, column=1)
            if not self._safe_set_hyperlink_formula(cell_uri, uri, uri):
                cell_uri.value = uri if uri else ''

            ws_props.cell(row=row_idx, column=2).value = label if label else ''
            ws_props.cell(row=row_idx, column=3).value = property_counts.get(label, 0)

        # Set column widths
        ws_props.column_dimensions['A'].width = 80
        ws_props.column_dimensions['B'].width = 30
        ws_props.column_dimensions['C'].width = 15

    def _print_export_summary(self, output_file: Path):
        """Print export summary statistics."""
        headers_count = len(set(k for s in self.subjects_data for k in s.keys())) - 1

        print(f"\nExcel file created: {output_file}")
        print(f"  - Sheet 'RDF Data': {len(self.subjects_data):,} rows, "
              f"{headers_count} columns")
        print(f"  - Sheet 'Property Mappings': {len(self.property_labels)} "
              f"property definitions")
        print(f"  - Using =HYPERLINK() formulas (unlimited links, no Excel corruption!)")

        # Count headers with external matches
        headers_with_links = sum(1 for uri in self.label_to_property_uri.values()
                                if uri in self.uri_to_exact_match or uri in self.uri_to_close_match)
        print(f"  - Column headers with external links: {headers_with_links}/{len(self.label_to_property_uri)}")


    def export_to_markdown(self, output_file: Path):
        """Export metadata and reference graph to Markdown format."""
        print(f"Exporting to Markdown: {output_file}")

        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                self._write_markdown_header(f)
                self._write_markdown_statistics(f)
                self._write_markdown_dcat(f)
                self._write_markdown_publications(f)
                self._write_markdown_namespaces(f)
                self._write_markdown_properties(f)

            print(f"Markdown file created: {output_file}")

        except Exception as e:
            print(f"Error exporting to Markdown: {e}")

    def _write_markdown_header(self, f):
        """Write markdown file header."""
        f.write("# RDF Metadata Export\n\n")
        source_name = self.input_file.name if self.input_file else "Session Data"
        f.write(f"**Source:** {source_name}\n\n")
        f.write(f"**Generated:** {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

    def _write_markdown_statistics(self, f):
        """Write dataset statistics section."""
        f.write("## Dataset Statistics\n\n")
        f.write(f"- **Total Triples:** {len(self.graph):,}\n")
        f.write(f"- **Total Subjects:** {len(self.subjects_data):,}\n")
        f.write(f"- **Total Properties:** {len(self.property_labels):,}\n\n")

    def _write_markdown_dcat(self, f):
        """Write DCAT metadata graph section."""
        f.write("## DCAT Metadata Graph\n\n")

        # Find the DCAT metadata graph by suffix
        dcat_graph_uri = next((g for g in self.named_graphs_data.keys() if g and g.endswith(DCAT_METADATA_SUFFIX)), None)

        if dcat_graph_uri:
            self._write_graph_section(f, dcat_graph_uri)
        else:
            f.write("*No DCAT metadata found*\n\n")

    def _write_markdown_publications(self, f):
        """Write publication reference graph section."""
        f.write("## Publication Reference Graph\n\n")

        # Find the publication reference graph by suffix
        pub_graph_uri = next((g for g in self.named_graphs_data.keys() if g and g.endswith(PUBLICATION_REFERENCE_SUFFIX)), None)

        if pub_graph_uri:
            self._write_graph_section(f, pub_graph_uri)
        else:
            f.write("*No publication reference data found*\n\n")

    def _write_graph_section(self, f, graph_uri: str):
        """Write a named graph section to markdown."""
        graph_subjects = defaultdict(list)
        for s, p, o in self.named_graphs_data[graph_uri]:
            graph_subjects[str(s)].append((p, o))

        for subject_uri in sorted(graph_subjects.keys()):
            # Find subject type
            subject_type = None
            for p, o in graph_subjects[subject_uri]:
                if str(p) == str(RDF.type):
                    subject_type = self._extract_local_name(str(o))
                    break

            f.write(f"### {subject_type if subject_type else 'Resource'}\n\n")

            for p, o in sorted(graph_subjects[subject_uri], key=lambda x: str(x[0])):
                if str(p) == str(RDF.type):
                    continue
                pred_label = self._extract_local_name(str(p))
                value_display = self._format_value_for_markdown(o)
                f.write(f"- **{pred_label}**: {value_display}\n")
            f.write("\n")

    def _write_markdown_namespaces(self, f):
        """Write namespace prefixes section."""
        f.write("## Namespace Prefixes\n\n")
        f.write("| Prefix | Namespace URI |\n")
        f.write("|--------|---------------|\n")

        for i, uri in enumerate(sorted(self.namespaces.keys())[:20]):
            escaped_uri = uri.replace('|', '\\|')
            f.write(f"| ns{i+1} | {escaped_uri} |\n")

        if len(self.namespaces) > 20:
            f.write(f"\n*...and {len(self.namespaces) - 20} more namespaces*\n")

    def _write_markdown_properties(self, f):
        """Write property reference graph section."""
        f.write("\n## Property Reference Graph\n\n")
        f.write("Complete mapping of all properties used in the dataset.\n\n")
        f.write("| Property URI | Label | Usage Count |\n")
        f.write("|--------------|-------|-------------|\n")

        # Get cached property usage counts
        property_counts = self._get_property_counts()

        for uri, label in sorted(self.property_labels.items(), key=lambda x: x[1]):
            escaped_uri = uri.replace('|', '\\|')
            escaped_label = label.replace('|', '\\|')
            count = property_counts.get(label, 0)
            f.write(f"| {escaped_uri} | {escaped_label} | {count} |\n")

        f.write("\n---\n\n")
        f.write("*For complete data records, please refer to the Excel file.*\n")

    def _format_value_for_markdown(self, value) -> str:
        """Format a value for markdown display."""
        value_str = str(value)
        if value_str in self.uri_to_label:
            return self.uri_to_label[value_str]
        return value_str


    def convert(self, output_dir: Path) -> bool:
        """
        Main conversion method.

        Args:
            output_dir: Directory for output files

        Returns:
            True if conversion succeeded, False otherwise
        """
        # Parse TriG file
        if not self.parse_trig():
            return False

        # Extract data
        self.extract_all_data()

        # Expand multi-valued properties
        self.expand_list_values()

        # Create output directory
        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate output filenames
        base_name = self.input_file.stem if self.input_file else "output"
        excel_file = output_dir / f"{base_name}.xlsx"
        markdown_file = output_dir / f"{base_name}.md"

        # Export to both formats
        self.export_to_excel(excel_file)
        self.export_to_markdown(markdown_file)

        print("\n✓ Conversion completed successfully!")
        return True



def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Convert TriG (RDF) files to Excel and Markdown formats',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python Transform_Trig_to_Excel.py data.trig
  python Transform_Trig_to_Excel.py data.trig --output-dir ./output
        """
    )

    parser.add_argument(
        'input_file',
        type=Path,
        help='Path to the input TriG file'
    )

    parser.add_argument(
        '--output-dir',
        type=Path,
        default=Path('./data'),
        help='Output directory for generated files (default: ./data)'
    )

    args = parser.parse_args()

    # Validate input file
    if not args.input_file.exists():
        print(f"Error: Input file not found: {args.input_file}")
        sys.exit(1)

    # Create converter and run
    converter = TriGConverter(args.input_file)
    success = converter.convert(args.output_dir)

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
